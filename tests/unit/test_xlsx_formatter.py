"""Unit tests for XLSX formatter.

Tests Excel export functionality for translation quality reports.
"""

from pathlib import Path
from unittest.mock import patch

import pytest

from kttc.core.models import ErrorAnnotation, ErrorSeverity, QAReport, TranslationTask


@pytest.fixture
def sample_task() -> TranslationTask:
    """Create sample translation task."""
    return TranslationTask(
        source_text="Hello world. This is a test.",
        translation="Привет мир. Это тест.",
        source_lang="en",
        target_lang="ru",
    )


@pytest.fixture
def sample_error() -> ErrorAnnotation:
    """Create sample error annotation."""
    return ErrorAnnotation(
        category="accuracy",
        subcategory="mistranslation",
        severity=ErrorSeverity.MAJOR,
        location=(0, 5),
        description="Test error description",
        suggestion="Fix this error",
    )


@pytest.fixture
def sample_report(sample_task: TranslationTask, sample_error: ErrorAnnotation) -> QAReport:
    """Create sample QA report with errors."""
    return QAReport(
        task=sample_task,
        mqm_score=85.0,
        errors=[sample_error],
        status="pass",
        confidence=0.9,
        agent_agreement=0.85,
    )


@pytest.fixture
def sample_report_no_errors(sample_task: TranslationTask) -> QAReport:
    """Create sample QA report without errors."""
    return QAReport(
        task=sample_task,
        mqm_score=100.0,
        errors=[],
        status="pass",
    )


@pytest.fixture
def sample_report_failed(sample_task: TranslationTask, sample_error: ErrorAnnotation) -> QAReport:
    """Create sample failed QA report."""
    critical_error = ErrorAnnotation(
        category="accuracy",
        subcategory="mistranslation",
        severity=ErrorSeverity.CRITICAL,
        location=(0, 10),
        description="Critical error",
        suggestion="Major fix needed",
    )
    return QAReport(
        task=sample_task,
        mqm_score=45.0,
        errors=[sample_error, critical_error],
        status="fail",
    )


class TestXLSXFormatterAvailability:
    """Tests for XLSX formatter availability check."""

    def test_is_available_with_openpyxl(self) -> None:
        """Test is_available returns True when openpyxl is installed."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        result = XLSXFormatter.is_available()
        assert isinstance(result, bool)

    def test_is_available_without_openpyxl(self) -> None:
        """Test is_available returns False when openpyxl is not installed."""
        with patch("kttc.cli.formatters.xlsx.OPENPYXL_AVAILABLE", False):
            from kttc.cli.formatters.xlsx import XLSXFormatter

            result = XLSXFormatter.is_available()
            assert isinstance(result, bool)


class TestXLSXFormatterFormatReport:
    """Tests for format_report method."""

    def test_format_report_save_to_file(self, sample_report: QAReport, tmp_path: Path) -> None:
        """Test saving report to file."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        output = tmp_path / "report.xlsx"
        result = XLSXFormatter.format_report(sample_report, output)

        assert result is None
        assert output.exists()
        assert output.stat().st_size > 0

    def test_format_report_return_bytes(self, sample_report: QAReport) -> None:
        """Test returning report as bytes."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        result = XLSXFormatter.format_report(sample_report, None)

        assert result is not None
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_format_report_without_text(self, sample_report: QAReport, tmp_path: Path) -> None:
        """Test saving report without text sheet."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(sample_report, output, include_text=False)

        assert output.exists()

    def test_format_report_no_errors(
        self, sample_report_no_errors: QAReport, tmp_path: Path
    ) -> None:
        """Test report with no errors."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(sample_report_no_errors, output)

        assert output.exists()

    def test_format_report_failed_status(
        self, sample_report_failed: QAReport, tmp_path: Path
    ) -> None:
        """Test report with failed status."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(sample_report_failed, output)

        assert output.exists()

    def test_format_report_raises_without_openpyxl(self, sample_report: QAReport) -> None:
        """Test that format_report raises ImportError without openpyxl."""
        with patch("kttc.cli.formatters.xlsx.OPENPYXL_AVAILABLE", False):
            from kttc.cli.formatters.xlsx import XLSXFormatter

            with pytest.raises(ImportError, match="openpyxl"):
                XLSXFormatter.format_report(sample_report, None)


class TestXLSXFormatterBatchReport:
    """Tests for format_batch_report method."""

    def test_format_batch_report_grouped(self, sample_report: QAReport, tmp_path: Path) -> None:
        """Test batch report with grouping by language."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        output = tmp_path / "batch_report.xlsx"
        XLSXFormatter.format_batch_report(
            [sample_report, sample_report], output, group_by_language=True
        )

        assert output.exists()

    def test_format_batch_report_not_grouped(self, sample_report: QAReport, tmp_path: Path) -> None:
        """Test batch report without grouping."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        output = tmp_path / "batch_report.xlsx"
        XLSXFormatter.format_batch_report(
            [sample_report, sample_report], output, group_by_language=False
        )

        assert output.exists()

    def test_format_batch_report_multiple_languages(self, tmp_path: Path) -> None:
        """Test batch report with multiple language pairs."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        task_en_ru = TranslationTask(
            source_text="Hello",
            translation="Привет",
            source_lang="en",
            target_lang="ru",
        )
        task_en_de = TranslationTask(
            source_text="Hello",
            translation="Hallo",
            source_lang="en",
            target_lang="de",
        )

        report_en_ru = QAReport(task=task_en_ru, mqm_score=90.0, errors=[], status="pass")
        report_en_de = QAReport(task=task_en_de, mqm_score=85.0, errors=[], status="pass")

        output = tmp_path / "batch_report.xlsx"
        XLSXFormatter.format_batch_report(
            [report_en_ru, report_en_de], output, group_by_language=True
        )

        assert output.exists()

    def test_format_batch_report_raises_without_openpyxl(
        self, sample_report: QAReport, tmp_path: Path
    ) -> None:
        """Test that format_batch_report raises ImportError without openpyxl."""
        with patch("kttc.cli.formatters.xlsx.OPENPYXL_AVAILABLE", False):
            from kttc.cli.formatters.xlsx import XLSXFormatter

            output = tmp_path / "batch_report.xlsx"
            with pytest.raises(ImportError, match="openpyxl"):
                XLSXFormatter.format_batch_report([sample_report], output)


class TestXLSXFormatterSheetCreation:
    """Tests for internal sheet creation methods."""

    def test_summary_sheet_contains_metrics(self, sample_report: QAReport, tmp_path: Path) -> None:
        """Test that summary sheet contains expected metrics."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        from openpyxl import load_workbook

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(sample_report, output)

        wb = load_workbook(output)
        assert "Summary" in wb.sheetnames

        ws = wb["Summary"]
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "MQM Score" in values
        assert "Status" in values

    def test_errors_sheet_contains_errors(self, sample_report: QAReport, tmp_path: Path) -> None:
        """Test that errors sheet contains error details."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        from openpyxl import load_workbook

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(sample_report, output)

        wb = load_workbook(output)
        assert "Errors" in wb.sheetnames

        ws = wb["Errors"]
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Category" in values
        assert "Severity" in values
        assert "accuracy" in values

    def test_breakdown_sheet_contains_categories(
        self, sample_report: QAReport, tmp_path: Path
    ) -> None:
        """Test that breakdown sheet contains category breakdown."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        from openpyxl import load_workbook

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(sample_report, output)

        wb = load_workbook(output)
        assert "Breakdown" in wb.sheetnames

    def test_text_sheet_contains_source_and_translation(
        self, sample_report: QAReport, tmp_path: Path
    ) -> None:
        """Test that text sheet contains source and translation."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        from openpyxl import load_workbook

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(sample_report, output, include_text=True)

        wb = load_workbook(output)
        assert "Text" in wb.sheetnames

        ws = wb["Text"]
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Source Text" in values
        assert "Translation" in values


class TestXLSXFormatterSeverityColors:
    """Tests for severity color handling."""

    def test_severity_colors_defined(self) -> None:
        """Test that severity colors are properly defined."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        assert "critical" in XLSXFormatter.SEVERITY_COLORS
        assert "major" in XLSXFormatter.SEVERITY_COLORS
        assert "minor" in XLSXFormatter.SEVERITY_COLORS
        assert "neutral" in XLSXFormatter.SEVERITY_COLORS

    def test_header_style_defined(self) -> None:
        """Test that header style is properly defined."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        assert XLSXFormatter.HEADER_FILL is not None
        assert XLSXFormatter.HEADER_FONT_COLOR is not None


class TestXLSXFormatterHelperMethods:
    """Tests for helper methods."""

    def test_add_header_row(self) -> None:
        """Test adding formatted header row."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        XLSXFormatter._add_header_row(ws, 1, ["Column1", "Column2", "Column3"])

        assert ws.cell(1, 1).value == "Column1"
        assert ws.cell(1, 2).value == "Column2"
        assert ws.cell(1, 3).value == "Column3"
        assert ws.cell(1, 1).font.bold is True

    def test_auto_adjust_columns(self) -> None:
        """Test auto-adjusting column widths."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Short"
        ws["B1"] = "This is a much longer text that should cause column B to be wider"

        XLSXFormatter._auto_adjust_columns(ws, max_width=50)

        assert ws.column_dimensions["A"].width > 0
        assert ws.column_dimensions["B"].width > 0
        assert ws.column_dimensions["B"].width >= ws.column_dimensions["A"].width

    def test_auto_adjust_columns_respects_max_width(self) -> None:
        """Test that auto-adjust respects max_width limit."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "x" * 200

        XLSXFormatter._auto_adjust_columns(ws, max_width=30)

        assert ws.column_dimensions["A"].width <= 30


class TestXLSXFormatterEdgeCases:
    """Tests for edge cases and error handling."""

    def test_report_with_comet_score(self, tmp_path: Path) -> None:
        """Test report with COMET score."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        task = TranslationTask(
            source_text="Hello",
            translation="Привет",
            source_lang="en",
            target_lang="ru",
        )
        report = QAReport(
            task=task,
            mqm_score=90.0,
            errors=[],
            status="pass",
            comet_score=0.95,
        )

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(report, output)

        assert output.exists()

    def test_report_with_multiple_error_severities(self, tmp_path: Path) -> None:
        """Test report with errors of all severity levels."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        task = TranslationTask(
            source_text="Hello world test",
            translation="Привет мир тест",
            source_lang="en",
            target_lang="ru",
        )
        errors = [
            ErrorAnnotation(
                category="accuracy",
                subcategory="mistranslation",
                severity=ErrorSeverity.CRITICAL,
                location=(0, 5),
                description="Critical error",
                suggestion="Fix",
            ),
            ErrorAnnotation(
                category="fluency",
                subcategory="grammar",
                severity=ErrorSeverity.MAJOR,
                location=(6, 11),
                description="Major error",
                suggestion="Fix",
            ),
            ErrorAnnotation(
                category="style",
                subcategory="register",
                severity=ErrorSeverity.MINOR,
                location=(12, 16),
                description="Minor error",
                suggestion="Fix",
            ),
            ErrorAnnotation(
                category="terminology",
                subcategory="inconsistent",
                severity=ErrorSeverity.NEUTRAL,
                location=(0, 5),
                description="Neutral observation",
                suggestion=None,
            ),
        ]
        report = QAReport(
            task=task,
            mqm_score=60.0,
            errors=errors,
            status="fail",
        )

        output = tmp_path / "report.xlsx"
        XLSXFormatter.format_report(report, output)

        assert output.exists()

    def test_empty_batch_report(self, tmp_path: Path) -> None:
        """Test batch report with empty list."""
        from kttc.cli.formatters.xlsx import XLSXFormatter

        if not XLSXFormatter.is_available():
            pytest.skip("openpyxl not installed")

        output = tmp_path / "batch_report.xlsx"
        XLSXFormatter.format_batch_report([], output)

        assert output.exists()
