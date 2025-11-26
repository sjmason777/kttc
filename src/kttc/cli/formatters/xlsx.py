# Copyright 2025 KTTC AI (https://github.com/kttc-ai)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""XLSX formatter for translation quality reports.

Provides Excel export functionality for enterprise reporting with:
- Summary scorecard sheet
- Detailed errors sheet
- Error breakdown by category
- Suggested corrections
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from kttc.core.models import QAReport

logger = logging.getLogger(__name__)

# Check if openpyxl is available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None


class XLSXFormatter:
    """Format QA reports as Excel spreadsheets.

    Provides enterprise-ready Excel export with multiple sheets:
    - Summary: Overall scores and metadata
    - Errors: Detailed error list with all attributes
    - Breakdown: Error counts by category and severity

    Example:
        >>> formatter = XLSXFormatter()
        >>> formatter.format_report(report, "output.xlsx")
    """

    # Color scheme for severity levels
    SEVERITY_COLORS = {
        "critical": "FF0000",  # Red
        "major": "FFA500",  # Orange
        "minor": "FFFF00",  # Yellow
        "neutral": "90EE90",  # Light green
    }

    # Header style
    HEADER_FILL = "4472C4"  # Blue
    HEADER_FONT_COLOR = "FFFFFF"  # White

    @staticmethod
    def is_available() -> bool:
        """Check if XLSX export is available (openpyxl installed).

        Returns:
            True if openpyxl is installed, False otherwise
        """
        return OPENPYXL_AVAILABLE

    @classmethod
    def format_report(
        cls,
        report: QAReport,
        output_path: str | Path | None = None,
        include_text: bool = True,
    ) -> bytes | None:
        """Format a QA report as Excel workbook.

        Args:
            report: QA report to format
            output_path: Optional path to save the workbook
            include_text: Whether to include source/translation text (default: True)

        Returns:
            Workbook bytes if no output_path, None if saved to file

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX export. " "Install it with: pip install openpyxl"
            )

        wb = Workbook()

        # Create sheets
        cls._create_summary_sheet(wb, report)
        cls._create_errors_sheet(wb, report)
        cls._create_breakdown_sheet(wb, report)

        if include_text:
            cls._create_text_sheet(wb, report)

        # Remove default sheet if it exists and is empty
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
            del wb["Sheet"]

        # Save or return bytes
        if output_path:
            wb.save(output_path)
            logger.info(f"XLSX report saved to {output_path}")
            return None
        else:
            from io import BytesIO

            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

    @classmethod
    def format_batch_report(
        cls,
        reports: list[QAReport],
        output_path: str | Path,
        group_by_language: bool = True,
    ) -> None:
        """Format multiple QA reports into a single Excel workbook.

        Args:
            reports: List of QA reports to format
            output_path: Path to save the workbook
            group_by_language: Create separate sheets per language pair (default: True)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX export. " "Install it with: pip install openpyxl"
            )

        wb = Workbook()

        # Create batch summary sheet
        cls._create_batch_summary_sheet(wb, reports)

        if group_by_language:
            # Group reports by language pair
            lang_groups: dict[str, list[QAReport]] = {}
            for report in reports:
                key = f"{report.task.source_lang}_{report.task.target_lang}"
                if key not in lang_groups:
                    lang_groups[key] = []
                lang_groups[key].append(report)

            # Create sheet for each language pair
            for lang_pair, lang_reports in lang_groups.items():
                cls._create_language_sheet(wb, lang_pair, lang_reports)
        else:
            # Single errors sheet for all reports
            cls._create_all_errors_sheet(wb, reports)

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)
        logger.info(f"Batch XLSX report saved to {output_path}")

    @classmethod
    def _create_summary_sheet(cls, wb: Any, report: QAReport) -> None:
        """Create summary scorecard sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Translation Quality Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # Generated timestamp
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(italic=True, color="666666")

        # Score card
        row = 4
        cls._add_header_row(ws, row, ["Metric", "Value"])
        row += 1

        metrics = [
            ("MQM Score", f"{report.mqm_score:.2f}"),
            ("Status", report.status.upper()),
            ("Source Language", report.task.source_lang.upper()),
            ("Target Language", report.task.target_lang.upper()),
            ("Word Count", str(report.task.word_count)),
            ("Total Errors", str(len(report.errors))),
        ]

        # Add optional metrics
        if report.confidence is not None:
            metrics.append(("Confidence", f"{report.confidence:.2f}"))
        if report.agent_agreement is not None:
            metrics.append(("Agent Agreement", f"{report.agent_agreement:.2f}"))
        if report.comet_score is not None:
            metrics.append(("COMET Score", f"{report.comet_score:.4f}"))

        for metric, value in metrics:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value

            # Color status cell
            if metric == "Status":
                fill_color = "90EE90" if value == "PASS" else "FF6B6B"
                ws[f"B{row}"].fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
            row += 1

        # Error severity breakdown
        row += 2
        ws[f"A{row}"] = "Error Severity Breakdown"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        severity_counts = {"critical": 0, "major": 0, "minor": 0, "neutral": 0}
        for error in report.errors:
            severity_counts[error.severity.value] = severity_counts.get(error.severity.value, 0) + 1

        cls._add_header_row(ws, row, ["Severity", "Count"])
        row += 1

        for severity, count in severity_counts.items():
            ws[f"A{row}"] = severity.capitalize()
            ws[f"B{row}"] = count
            if count > 0:
                fill_color = cls.SEVERITY_COLORS.get(severity, "FFFFFF")
                ws[f"A{row}"].fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
            row += 1

        # Auto-adjust column widths
        cls._auto_adjust_columns(ws)

    @classmethod
    def _create_errors_sheet(cls, wb: Any, report: QAReport) -> None:
        """Create detailed errors sheet."""
        ws = wb.create_sheet("Errors")

        if not report.errors:
            ws["A1"] = "No errors found"
            ws["A1"].font = Font(italic=True, color="28a745")
            return

        # Headers
        headers = [
            "Category",
            "Subcategory",
            "Severity",
            "Location Start",
            "Location End",
            "Description",
            "Suggestion",
        ]
        cls._add_header_row(ws, 1, headers)

        # Error data
        for row_idx, error in enumerate(report.errors, start=2):
            ws[f"A{row_idx}"] = error.category
            ws[f"B{row_idx}"] = error.subcategory
            ws[f"C{row_idx}"] = error.severity.value.capitalize()
            ws[f"D{row_idx}"] = error.location[0]
            ws[f"E{row_idx}"] = error.location[1]
            ws[f"F{row_idx}"] = error.description
            ws[f"G{row_idx}"] = error.suggestion or ""

            # Color severity cell
            fill_color = cls.SEVERITY_COLORS.get(error.severity.value, "FFFFFF")
            ws[f"C{row_idx}"].fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

        # Auto-adjust column widths
        cls._auto_adjust_columns(ws, max_width=50)

    @classmethod
    def _create_breakdown_sheet(cls, wb: Any, report: QAReport) -> None:
        """Create error breakdown by category sheet."""
        ws = wb.create_sheet("Breakdown")

        # Category breakdown
        ws["A1"] = "Error Breakdown by Category"
        ws["A1"].font = Font(size=14, bold=True)

        # Count errors by category
        category_counts: dict[str, dict[str, int]] = {}
        for error in report.errors:
            if error.category not in category_counts:
                category_counts[error.category] = {
                    "critical": 0,
                    "major": 0,
                    "minor": 0,
                    "neutral": 0,
                    "total": 0,
                }
            category_counts[error.category][error.severity.value] += 1
            category_counts[error.category]["total"] += 1

        if not category_counts:
            ws["A3"] = "No errors to analyze"
            return

        # Headers
        headers = ["Category", "Critical", "Major", "Minor", "Neutral", "Total"]
        cls._add_header_row(ws, 3, headers)

        # Data
        row = 4
        for category, counts in sorted(category_counts.items()):
            ws[f"A{row}"] = category.capitalize()
            ws[f"B{row}"] = counts["critical"]
            ws[f"C{row}"] = counts["major"]
            ws[f"D{row}"] = counts["minor"]
            ws[f"E{row}"] = counts["neutral"]
            ws[f"F{row}"] = counts["total"]

            # Color severity cells
            for col, sev in [("B", "critical"), ("C", "major"), ("D", "minor")]:
                if counts[sev] > 0:
                    fill_color = cls.SEVERITY_COLORS[sev]
                    ws[f"{col}{row}"].fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
            row += 1

        # Totals row
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = sum(c["critical"] for c in category_counts.values())
        ws[f"C{row}"] = sum(c["major"] for c in category_counts.values())
        ws[f"D{row}"] = sum(c["minor"] for c in category_counts.values())
        ws[f"E{row}"] = sum(c["neutral"] for c in category_counts.values())
        ws[f"F{row}"] = sum(c["total"] for c in category_counts.values())

        cls._auto_adjust_columns(ws)

    @classmethod
    def _create_text_sheet(cls, wb: Any, report: QAReport) -> None:
        """Create sheet with source and translation text."""
        ws = wb.create_sheet("Text")

        ws["A1"] = "Source Text"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = report.task.source_text
        ws["A2"].alignment = Alignment(wrap_text=True)

        ws["A4"] = "Translation"
        ws["A4"].font = Font(bold=True)
        ws["A5"] = report.task.translation
        ws["A5"].alignment = Alignment(wrap_text=True)

        # Set column width
        ws.column_dimensions["A"].width = 100

    @classmethod
    def _create_batch_summary_sheet(cls, wb: Any, reports: list[QAReport]) -> None:
        """Create batch summary sheet for multiple reports."""
        ws = wb.create_sheet("Batch Summary", 0)

        ws["A1"] = "Batch Translation Quality Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:F1")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(italic=True, color="666666")

        # Overall statistics
        row = 4
        total_reports = len(reports)
        passed = sum(1 for r in reports if r.status == "pass")
        failed = total_reports - passed
        avg_score = sum(r.mqm_score for r in reports) / total_reports if reports else 0
        total_errors = sum(len(r.errors) for r in reports)

        ws[f"A{row}"] = "Batch Statistics"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        stats = [
            ("Total Reports", total_reports),
            ("Passed", passed),
            ("Failed", failed),
            ("Pass Rate", f"{(passed / total_reports * 100):.1f}%" if total_reports else "N/A"),
            ("Average MQM Score", f"{avg_score:.2f}"),
            ("Total Errors", total_errors),
        ]

        for stat_name, stat_value in stats:
            ws[f"A{row}"] = stat_name
            ws[f"B{row}"] = stat_value
            row += 1

        # Individual reports table
        row += 2
        headers = ["#", "Source Lang", "Target Lang", "MQM Score", "Status", "Errors"]
        cls._add_header_row(ws, row, headers)
        row += 1

        for idx, report in enumerate(reports, start=1):
            ws[f"A{row}"] = idx
            ws[f"B{row}"] = report.task.source_lang.upper()
            ws[f"C{row}"] = report.task.target_lang.upper()
            ws[f"D{row}"] = f"{report.mqm_score:.2f}"
            ws[f"E{row}"] = report.status.upper()
            ws[f"F{row}"] = len(report.errors)

            # Color status
            fill_color = "90EE90" if report.status == "pass" else "FF6B6B"
            ws[f"E{row}"].fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            row += 1

        cls._auto_adjust_columns(ws)

    @classmethod
    def _create_language_sheet(cls, wb: Any, lang_pair: str, reports: list[QAReport]) -> None:
        """Create sheet for specific language pair."""
        # Sanitize sheet name (Excel has 31 char limit)
        sheet_name = lang_pair[:31]
        ws = wb.create_sheet(sheet_name)

        ws["A1"] = f"Errors for {lang_pair.replace('_', ' -> ')}"
        ws["A1"].font = Font(size=14, bold=True)

        # Headers
        headers = [
            "Report #",
            "Category",
            "Subcategory",
            "Severity",
            "Description",
            "Suggestion",
        ]
        cls._add_header_row(ws, 3, headers)

        row = 4
        for report_idx, report in enumerate(reports, start=1):
            for error in report.errors:
                ws[f"A{row}"] = report_idx
                ws[f"B{row}"] = error.category
                ws[f"C{row}"] = error.subcategory
                ws[f"D{row}"] = error.severity.value.capitalize()
                ws[f"E{row}"] = error.description
                ws[f"F{row}"] = error.suggestion or ""

                fill_color = cls.SEVERITY_COLORS.get(error.severity.value, "FFFFFF")
                ws[f"D{row}"].fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
                row += 1

        cls._auto_adjust_columns(ws, max_width=50)

    @classmethod
    def _create_all_errors_sheet(cls, wb: Any, reports: list[QAReport]) -> None:
        """Create single sheet with all errors from all reports."""
        ws = wb.create_sheet("All Errors")

        headers = [
            "Report #",
            "Source Lang",
            "Target Lang",
            "Category",
            "Subcategory",
            "Severity",
            "Description",
        ]
        cls._add_header_row(ws, 1, headers)

        row = 2
        for report_idx, report in enumerate(reports, start=1):
            for error in report.errors:
                ws[f"A{row}"] = report_idx
                ws[f"B{row}"] = report.task.source_lang
                ws[f"C{row}"] = report.task.target_lang
                ws[f"D{row}"] = error.category
                ws[f"E{row}"] = error.subcategory
                ws[f"F{row}"] = error.severity.value.capitalize()
                ws[f"G{row}"] = error.description

                fill_color = cls.SEVERITY_COLORS.get(error.severity.value, "FFFFFF")
                ws[f"F{row}"].fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
                row += 1

        cls._auto_adjust_columns(ws, max_width=50)

    @classmethod
    def _add_header_row(cls, ws: Any, row: int, headers: list[str]) -> None:
        """Add formatted header row to worksheet."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color=cls.HEADER_FONT_COLOR)
            cell.fill = PatternFill(
                start_color=cls.HEADER_FILL,
                end_color=cls.HEADER_FILL,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")

    @classmethod
    def _auto_adjust_columns(cls, ws: Any, max_width: int = 30) -> None:
        """Auto-adjust column widths based on content."""
        try:
            for column_cells in ws.columns:
                # Convert generator to tuple to allow indexing and iteration
                column_cells_tuple = tuple(column_cells)
                if not column_cells_tuple:
                    continue
                max_length = 0
                first_cell = column_cells_tuple[0]
                # Handle both Cell objects and tuples
                if hasattr(first_cell, "column_letter"):
                    column = first_cell.column_letter
                elif hasattr(first_cell, "column"):
                    from openpyxl.utils import get_column_letter

                    column = get_column_letter(first_cell.column)
                else:
                    continue
                for cell in column_cells_tuple:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except (AttributeError, TypeError):
                        pass
                adjusted_width = min(max_length + 2, max_width)
                ws.column_dimensions[column].width = adjusted_width
        except Exception:
            logging.warning("Failed to auto-adjust columns", exc_info=True)
